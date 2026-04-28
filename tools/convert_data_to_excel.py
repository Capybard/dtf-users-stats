import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

input_path = Path("data/data.json")
output_path = Path("data/data.xlsx")

with input_path.open("r", encoding="utf-8") as f:
    data = json.load(f)

wb = Workbook()
wb.remove(wb.active)

header_fill = PatternFill("solid", fgColor="D9EAF7")
header_font = Font(bold=True)

def autosize(ws):
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 60)

def style_header(ws):
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

for section, value in data.items():
    ws = wb.create_sheet(title=section[:31])

    if isinstance(value, dict):
        ws.append(["Значение", "Количество"])
        for k, v in value.items():
            ws.append([k, v])

    elif isinstance(value, list):
        max_len = max(len(row) for row in value)
        ws.append([f"Колонка {i}" for i in range(1, max_len + 1)])
        for row in value:
            ws.append(row)

    else:
        ws.append(["Значение"])
        ws.append([value])

    style_header(ws)
    autosize(ws)
    ws.freeze_panes = "A2"

wb.save(output_path)
print(f"Готово: {output_path.resolve()}")